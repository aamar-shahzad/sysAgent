"""
Spreadsheet tool for SysAgent CLI - Create and manage Excel/CSV files.
"""

import os
import csv
import json
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

from .base import BaseTool, ToolMetadata, ToolResult, register_tool
from ..types import ToolCategory
from ..utils.platform import detect_platform, Platform


@register_tool
class SpreadsheetTool(BaseTool):
    """Tool for creating and managing spreadsheets (Excel/CSV)."""
    
    def _get_metadata(self) -> ToolMetadata:
        return ToolMetadata(
            name="spreadsheet_tool",
            description="Create, edit, and manage spreadsheets (Excel, CSV)",
            category=ToolCategory.FILE,
            permissions=["file_access"],
            version="1.0.0"
        )

    def _execute(self, action: str, **kwargs) -> ToolResult:
        try:
            actions = {
                "create": self._create_spreadsheet,
                "create_excel": self._create_excel,
                "read": self._read_spreadsheet,
                "write_row": self._write_row,
                "write_rows": self._write_rows,
                "add_sheet": self._add_sheet,
                "get_cell": self._get_cell,
                "set_cell": self._set_cell,
                "create_data_entry": self._create_data_entry_form,
                "create_template": self._create_template,
                "to_json": self._to_json,
                "from_json": self._from_json,
                "open": self._open_spreadsheet,
            }
            
            if action in actions:
                return actions[action](**kwargs)
            else:
                return ToolResult(
                    success=False,
                    data={},
                    message=f"Unknown action: {action}",
                    error=f"Supported actions: {list(actions.keys())}"
                )
        except Exception as e:
            return ToolResult(
                success=False,
                data={},
                message=f"Spreadsheet operation failed: {str(e)}",
                error=str(e)
            )

    def _create_spreadsheet(self, **kwargs) -> ToolResult:
        """Create a new CSV spreadsheet."""
        path = kwargs.get("path")
        headers = kwargs.get("headers", [])
        data = kwargs.get("data", [])
        title = kwargs.get("title", "")
        
        if not path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_title = "".join(c if c.isalnum() or c in " -_" else "" for c in (title or "spreadsheet"))
            path = str(Path.home() / "Documents" / f"{safe_title}_{timestamp}.csv")
        
        try:
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                if headers:
                    writer.writerow(headers)
                for row in data:
                    writer.writerow(row)
            
            return ToolResult(
                success=True,
                data={"path": path, "headers": headers, "rows": len(data)},
                message=f"Spreadsheet created: {path}"
            )
        except Exception as e:
            return ToolResult(
                success=False,
                data={},
                message=f"Failed to create spreadsheet: {str(e)}",
                error=str(e)
            )

    def _create_excel(self, **kwargs) -> ToolResult:
        """Create an Excel file (requires openpyxl)."""
        path = kwargs.get("path")
        headers = kwargs.get("headers", [])
        data = kwargs.get("data", [])
        title = kwargs.get("title", "Sheet1")
        sheets = kwargs.get("sheets", [])  # List of {name, headers, data}
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return ToolResult(
                success=False,
                data={},
                message="openpyxl not installed. Install with: pip install openpyxl",
                error="Missing dependency"
            )
        
        if not path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_title = "".join(c if c.isalnum() or c in " -_" else "" for c in (title or "spreadsheet"))
            path = str(Path.home() / "Documents" / f"{safe_title}_{timestamp}.xlsx")
        
        try:
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            
            wb = Workbook()
            ws = wb.active
            ws.title = title
            
            # Style for headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Write headers
            if headers:
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
            
            # Write data
            for row_idx, row in enumerate(data, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Create additional sheets
            for sheet_info in sheets:
                sheet_name = sheet_info.get("name", f"Sheet{len(wb.sheetnames)+1}")
                new_ws = wb.create_sheet(title=sheet_name)
                
                sheet_headers = sheet_info.get("headers", [])
                sheet_data = sheet_info.get("data", [])
                
                if sheet_headers:
                    for col, header in enumerate(sheet_headers, 1):
                        cell = new_ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                
                for row_idx, row in enumerate(sheet_data, 2):
                    for col_idx, value in enumerate(row, 1):
                        new_ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(path)
            
            return ToolResult(
                success=True,
                data={
                    "path": path,
                    "sheets": [title] + [s.get("name") for s in sheets],
                    "rows": len(data)
                },
                message=f"Excel file created: {path}"
            )
        except Exception as e:
            return ToolResult(
                success=False,
                data={},
                message=f"Failed to create Excel file: {str(e)}",
                error=str(e)
            )

    def _read_spreadsheet(self, **kwargs) -> ToolResult:
        """Read a spreadsheet (CSV or Excel)."""
        path = kwargs.get("path")
        sheet = kwargs.get("sheet")  # For Excel
        
        if not path:
            return ToolResult(
                success=False,
                data={},
                message="No file path provided",
                error="Missing path"
            )
        
        if not Path(path).exists():
            return ToolResult(
                success=False,
                data={},
                message=f"File not found: {path}",
                error="File not found"
            )
        
        try:
            ext = Path(path).suffix.lower()
            
            if ext == '.csv':
                with open(path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                
                return ToolResult(
                    success=True,
                    data={
                        "path": path,
                        "headers": rows[0] if rows else [],
                        "data": rows[1:] if len(rows) > 1 else [],
                        "total_rows": len(rows)
                    },
                    message=f"Read {len(rows)} rows from {path}"
                )
            
            elif ext in ['.xlsx', '.xls']:
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    return ToolResult(
                        success=False,
                        data={},
                        message="openpyxl not installed for Excel support",
                        error="Missing dependency"
                    )
                
                wb = load_workbook(path, data_only=True)
                ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
                
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(list(row))
                
                return ToolResult(
                    success=True,
                    data={
                        "path": path,
                        "sheet": ws.title,
                        "sheets": wb.sheetnames,
                        "headers": rows[0] if rows else [],
                        "data": rows[1:] if len(rows) > 1 else [],
                        "total_rows": len(rows)
                    },
                    message=f"Read {len(rows)} rows from {path}"
                )
            else:
                return ToolResult(
                    success=False,
                    data={},
                    message=f"Unsupported file format: {ext}",
                    error="Use .csv or .xlsx"
                )
        except Exception as e:
            return ToolResult(
                success=False,
                data={},
                message=f"Failed to read spreadsheet: {str(e)}",
                error=str(e)
            )

    def _write_row(self, **kwargs) -> ToolResult:
        """Append a row to a CSV file."""
        path = kwargs.get("path")
        row = kwargs.get("row", [])
        
        if not path or not row:
            return ToolResult(
                success=False,
                data={},
                message="Path and row data required",
                error="Missing parameters"
            )
        
        try:
            with open(path, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(row)
            
            return ToolResult(
                success=True,
                data={"path": path, "row": row},
                message=f"Row added to {path}"
            )
        except Exception as e:
            return ToolResult(
                success=False,
                data={},
                message=f"Failed to write row: {str(e)}",
                error=str(e)
            )

    def _write_rows(self, **kwargs) -> ToolResult:
        """Append multiple rows to a CSV file."""
        path = kwargs.get("path")
        rows = kwargs.get("rows", [])
        
        if not path or not rows:
            return ToolResult(
                success=False,
                data={},
                message="Path and rows data required",
                error="Missing parameters"
            )
        
        try:
            with open(path, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            
            return ToolResult(
                success=True,
                data={"path": path, "rows_added": len(rows)},
                message=f"Added {len(rows)} rows to {path}"
            )
        except Exception as e:
            return ToolResult(
                success=False,
                data={},
                message=f"Failed to write rows: {str(e)}",
                error=str(e)
            )

    def _add_sheet(self, **kwargs) -> ToolResult:
        """Add a new sheet to an Excel file."""
        path = kwargs.get("path")
        sheet_name = kwargs.get("name", "New Sheet")
        headers = kwargs.get("headers", [])
        
        if not path:
            return ToolResult(
                success=False,
                data={},
                message="Excel file path required",
                error="Missing path"
            )
        
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return ToolResult(
                success=False,
                data={},
                message="openpyxl not installed",
                error="Missing dependency"
            )
        
        try:
            wb = load_workbook(path)
            ws = wb.create_sheet(title=sheet_name)
            
            if headers:
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
            
            wb.save(path)
            
            return ToolResult(
                success=True,
                data={"path": path, "sheet": sheet_name, "sheets": wb.sheetnames},
                message=f"Added sheet '{sheet_name}' to {path}"
            )
        except Exception as e:
            return ToolResult(
                success=False,
                data={},
                message=f"Failed to add sheet: {str(e)}",
                error=str(e)
            )

    def _get_cell(self, **kwargs) -> ToolResult:
        """Get value of a specific cell."""
        path = kwargs.get("path")
        cell = kwargs.get("cell")  # e.g., "A1"
        sheet = kwargs.get("sheet")
        
        if not path or not cell:
            return ToolResult(
                success=False,
                data={},
                message="Path and cell reference required",
                error="Missing parameters"
            )
        
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ToolResult(
                success=False,
                data={},
                message="openpyxl not installed",
                error="Missing dependency"
            )
        
        try:
            wb = load_workbook(path, data_only=True)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            value = ws[cell].value
            
            return ToolResult(
                success=True,
                data={"cell": cell, "value": value, "sheet": ws.title},
                message=f"Cell {cell} = {value}"
            )
        except Exception as e:
            return ToolResult(
                success=False,
                data={},
                message=f"Failed to get cell: {str(e)}",
                error=str(e)
            )

    def _set_cell(self, **kwargs) -> ToolResult:
        """Set value of a specific cell."""
        path = kwargs.get("path")
        cell = kwargs.get("cell")
        value = kwargs.get("value")
        sheet = kwargs.get("sheet")
        
        if not path or not cell:
            return ToolResult(
                success=False,
                data={},
                message="Path and cell reference required",
                error="Missing parameters"
            )
        
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ToolResult(
                success=False,
                data={},
                message="openpyxl not installed",
                error="Missing dependency"
            )
        
        try:
            wb = load_workbook(path)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            ws[cell] = value
            wb.save(path)
            
            return ToolResult(
                success=True,
                data={"cell": cell, "value": value, "sheet": ws.title},
                message=f"Set {cell} = {value}"
            )
        except Exception as e:
            return ToolResult(
                success=False,
                data={},
                message=f"Failed to set cell: {str(e)}",
                error=str(e)
            )

    def _create_data_entry_form(self, **kwargs) -> ToolResult:
        """Create a data entry spreadsheet with validation and formatting."""
        path = kwargs.get("path")
        title = kwargs.get("title", "Data Entry Form")
        fields = kwargs.get("fields", [])  # List of field names/columns
        description = kwargs.get("description", "")
        
        if not fields:
            fields = ["ID", "Name", "Date", "Description", "Amount", "Status"]
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback to CSV
            if not path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                path = str(Path.home() / "Documents" / f"data_entry_{timestamp}.csv")
            
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(fields)
            
            return ToolResult(
                success=True,
                data={"path": path, "fields": fields, "format": "csv"},
                message=f"Created CSV data entry form: {path}"
            )
        
        if not path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_title = "".join(c if c.isalnum() or c in " -_" else "" for c in title)
            path = str(Path.home() / "Documents" / f"{safe_title}_{timestamp}.xlsx")
        
        try:
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Entry"
            
            # Styles
            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(bold=True, size=11, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
            ws.cell(row=1, column=1, value=title).font = title_font
            
            # Description
            if description:
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(fields))
                ws.cell(row=2, column=1, value=description)
            
            # Headers (row 4)
            for col, field in enumerate(fields, 1):
                cell = ws.cell(row=4, column=col, value=field)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # Auto-width
                ws.column_dimensions[get_column_letter(col)].width = max(12, len(field) + 4)
            
            # Add empty rows for data entry
            for row in range(5, 25):
                for col in range(1, len(fields) + 1):
                    ws.cell(row=row, column=col).border = thin_border
            
            wb.save(path)
            
            return ToolResult(
                success=True,
                data={"path": path, "title": title, "fields": fields, "format": "xlsx"},
                message=f"Created data entry form: {path}"
            )
        except Exception as e:
            return ToolResult(
                success=False,
                data={},
                message=f"Failed to create data entry form: {str(e)}",
                error=str(e)
            )

    def _create_template(self, **kwargs) -> ToolResult:
        """Create a spreadsheet from a template."""
        template = kwargs.get("template")  # budget, inventory, timesheet, contacts
        path = kwargs.get("path")
        title = kwargs.get("title", "")
        
        templates = {
            "budget": {
                "title": "Budget Tracker",
                "headers": ["Category", "Description", "Budgeted", "Actual", "Difference", "Notes"],
                "data": [
                    ["Income", "Salary", 5000, 5000, 0, ""],
                    ["Income", "Other", 500, 0, -500, ""],
                    ["Housing", "Rent/Mortgage", 1500, 1500, 0, ""],
                    ["Utilities", "Electric", 150, 0, -150, ""],
                    ["Utilities", "Water", 50, 0, -50, ""],
                    ["Food", "Groceries", 400, 0, -400, ""],
                    ["Transport", "Gas/Fuel", 200, 0, -200, ""],
                ]
            },
            "inventory": {
                "title": "Inventory List",
                "headers": ["Item ID", "Item Name", "Category", "Quantity", "Unit Price", "Total Value", "Location", "Reorder Level"],
                "data": []
            },
            "timesheet": {
                "title": "Weekly Timesheet",
                "headers": ["Date", "Project", "Task", "Start Time", "End Time", "Hours", "Notes"],
                "data": []
            },
            "contacts": {
                "title": "Contact List",
                "headers": ["Name", "Email", "Phone", "Company", "Title", "Address", "Notes"],
                "data": []
            },
            "expenses": {
                "title": "Expense Tracker",
                "headers": ["Date", "Category", "Description", "Amount", "Payment Method", "Receipt", "Status"],
                "data": []
            }
        }
        
        if template not in templates:
            return ToolResult(
                success=False,
                data={"available_templates": list(templates.keys())},
                message=f"Unknown template: {template}",
                error=f"Available templates: {list(templates.keys())}"
            )
        
        tmpl = templates[template]
        
        return self._create_excel(
            path=path,
            title=title or tmpl["title"],
            headers=tmpl["headers"],
            data=tmpl["data"]
        )

    def _to_json(self, **kwargs) -> ToolResult:
        """Convert spreadsheet to JSON."""
        path = kwargs.get("path")
        output = kwargs.get("output")
        
        if not path:
            return ToolResult(
                success=False,
                data={},
                message="Input path required",
                error="Missing path"
            )
        
        result = self._read_spreadsheet(path=path)
        if not result.success:
            return result
        
        headers = result.data.get("headers", [])
        rows = result.data.get("data", [])
        
        # Convert to list of dicts
        json_data = []
        for row in rows:
            row_dict = {}
            for i, header in enumerate(headers):
                row_dict[header] = row[i] if i < len(row) else None
            json_data.append(row_dict)
        
        if output:
            try:
                with open(output, 'w', encoding='utf-8') as f:
                    json.dump(json_data, f, indent=2, default=str)
                
                return ToolResult(
                    success=True,
                    data={"path": output, "records": len(json_data)},
                    message=f"Exported {len(json_data)} records to {output}"
                )
            except Exception as e:
                return ToolResult(
                    success=False,
                    data={},
                    message=f"Failed to export: {str(e)}",
                    error=str(e)
                )
        else:
            return ToolResult(
                success=True,
                data={"json": json_data, "records": len(json_data)},
                message=f"Converted {len(json_data)} records"
            )

    def _from_json(self, **kwargs) -> ToolResult:
        """Create spreadsheet from JSON data."""
        json_path = kwargs.get("json_path")
        json_data = kwargs.get("json_data")
        output = kwargs.get("output")
        
        if json_path:
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
            except Exception as e:
                return ToolResult(
                    success=False,
                    data={},
                    message=f"Failed to read JSON: {str(e)}",
                    error=str(e)
                )
        
        if not json_data or not isinstance(json_data, list):
            return ToolResult(
                success=False,
                data={},
                message="JSON data must be a list of objects",
                error="Invalid JSON format"
            )
        
        # Extract headers from first record
        headers = list(json_data[0].keys()) if json_data else []
        
        # Convert to rows
        data = []
        for record in json_data:
            row = [record.get(h, "") for h in headers]
            data.append(row)
        
        if not output:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output = str(Path.home() / "Documents" / f"from_json_{timestamp}.csv")
        
        return self._create_spreadsheet(path=output, headers=headers, data=data)

    def _open_spreadsheet(self, **kwargs) -> ToolResult:
        """Open spreadsheet in default application."""
        path = kwargs.get("path")
        app = kwargs.get("app")
        
        if not path:
            return ToolResult(
                success=False,
                data={},
                message="File path required",
                error="Missing path"
            )
        
        if not Path(path).exists():
            return ToolResult(
                success=False,
                data={},
                message=f"File not found: {path}",
                error="File not found"
            )
        
        try:
            import subprocess
            platform = detect_platform()
            
            if platform == Platform.MACOS:
                if app:
                    subprocess.Popen(["open", "-a", app, path])
                else:
                    subprocess.Popen(["open", path])
            elif platform == Platform.WINDOWS:
                if app:
                    subprocess.Popen([app, path])
                else:
                    os.startfile(path)
            else:  # Linux
                if app:
                    subprocess.Popen([app, path])
                else:
                    subprocess.Popen(["xdg-open", path])
            
            return ToolResult(
                success=True,
                data={"path": path},
                message=f"Opened: {path}"
            )
        except Exception as e:
            return ToolResult(
                success=False,
                data={},
                message=f"Failed to open spreadsheet: {str(e)}",
                error=str(e)
            )

    def get_usage_examples(self) -> List[str]:
        return [
            "Create CSV: spreadsheet_tool --action create --headers 'Name,Email,Phone'",
            "Create Excel: spreadsheet_tool --action create_excel --headers 'ID,Name,Amount'",
            "Create data entry form: spreadsheet_tool --action create_data_entry --fields 'Name,Date,Amount'",
            "Create from template: spreadsheet_tool --action create_template --template budget",
            "Read spreadsheet: spreadsheet_tool --action read --path '/path/to/file.csv'",
            "Add row: spreadsheet_tool --action write_row --path '/path/to/file.csv' --row 'John,john@email.com'",
        ]
